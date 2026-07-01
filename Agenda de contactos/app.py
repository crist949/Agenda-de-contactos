
import os, re
from flask import (Flask, render_template, request,
                   redirect, url_for, session, flash, jsonify)
import openpyxl
from openpyxl import Workbook

# ── Configuración ──────────────────────────────────────────
app = Flask(__name__)
app.secret_key = 'clave_secreta_proyecto_2026'

USUARIO_CORRECTO  = 'admin'
CONTRASENA_CORRECTA = '1234'

# Ruta del Excel 
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'contactos.xlsx')

CATEGORIAS = ['Familia', 'Trabajo', 'Amigos', 'Otro']
ENCABEZADOS = ['Nombre', 'Apellido', 'Teléfono', 'Correo',
               'Dirección', 'Categoría', 'Favorito']

# ── openpyxl Funcionamiento ───────────────────────────────────────


#crea el archivo
def _init_excel():
    
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(ENCABEZADOS)

        Contactos = [
            ['Sharon',   'Carranza',    '88001122', 'carranza@gmail.com',      'San José',    'Familia',  'Sí'],
            ['Santiago',  'Fallas',     '77334455', 'fallassanti@gmail.com',    'Heredia',     'Trabajo',  'No'],
            ['Matias',   'Sequeira',     '66778899', 'sequeira@gmail.com',    'Cartago',     'Amigos',   'Sí'],
            ['Eliberto',    'Vazquez', '55443322', 'eli@hotmail.com',   'Alajuela',    'Otro',     'No'],
            ['Kendall',   'Espinoza',  '44556677', 'kendallesp@gmail.com',    'Liberia',     'Familia',  'Sí'],
        ]
        for row in  Contactos:
            ws.append(row)
        wb.save(EXCEL_PATH)


def _leer_contactos():
    
    _init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    contactos = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(row):          
            contactos.append({
                'fila':      i,
                'nombre':    row[0] or '',
                'apellido':  row[1] or '',
                'telefono':  row[2] or '',
                'correo':    row[3] or '',
                'direccion': row[4] or '',
                'categoria': row[5] or 'Otro',
                'favorito':  row[6] or 'No',
            })
    return contactos


#agrega nuevas filas
def _guardar_contacto(datos):

    _init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([
        datos['nombre'], datos['apellido'], datos['telefono'],
        datos['correo'], datos['direccion'], datos['categoria'],
        datos['favorito'],
    ])
    wb.save(EXCEL_PATH)


#modifica filas exactas
def _actualizar_contacto(fila, datos):
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.cell(fila, 1, datos['nombre'])
    ws.cell(fila, 2, datos['apellido'])
    ws.cell(fila, 3, datos['telefono'])
    ws.cell(fila, 4, datos['correo'])
    ws.cell(fila, 5, datos['direccion'])
    ws.cell(fila, 6, datos['categoria'])
    ws.cell(fila, 7, datos['favorito'])
    wb.save(EXCEL_PATH)


#elimina filas
def _eliminar_contacto(fila):
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.delete_rows(fila)
    wb.save(EXCEL_PATH)


#verifica lista de errores
def _validar(nombre, telefono, correo):
    
    errores = []
    if not nombre.strip():
        errores.append('El nombre es obligatorio.')
    if not re.fullmatch(r'\d{8}', telefono.strip()):
        errores.append('El teléfono debe tener exactamente 8 dígitos.')
    if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', correo.strip()):
        errores.append('El correo electrónico no tiene un formato válido.')
    return errores


#rediriqe el login
def login_required(f):
    
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# / RUTAS /


# ── Inicio ────────────────────────────────────────
@app.route('/')
def inicio():
    return render_template('inicio.html')


# ── Login ──────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        usuario   = request.form.get('usuario', '')
        contrasena = request.form.get('contrasena', '')
        if usuario == USUARIO_CORRECTO and contrasena == CONTRASENA_CORRECTA:
            session['logged_in'] = True
            return redirect(url_for('principal'))
        else:
            error = 'Usuario o Contraseña incorrecta'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Principal ─────────────────────────
@app.route('/principal', methods=['GET'])
@login_required
def principal():
    contactos  = _leer_contactos()
    busqueda   = request.args.get('busqueda', '').strip()
    ordenar    = request.args.get('ordenar', '')

    if busqueda:
        contactos = [c for c in contactos
                     if busqueda.lower() in c['nombre'].lower()
                     or busqueda.lower() in c['apellido'].lower()]

    if ordenar == 'az':
        contactos = sorted(contactos, key=lambda c: c['nombre'].lower())

    return render_template('index.html',
                           contactos=contactos,
                           busqueda=busqueda,
                           ordenar=ordenar)


# ── Agregar contacto ───────────────────────────────────────
@app.route('/agregar_usuario', methods=['GET', 'POST'])
@login_required
def agregar_usuario():
    errores = []
    if request.method == 'POST':
        datos = {
            'nombre':    request.form.get('nombre', '').strip(),
            'apellido':  request.form.get('apellido', '').strip(),
            'telefono':  request.form.get('telefono', '').strip(),
            'correo':    request.form.get('correo', '').strip(),
            'direccion': request.form.get('direccion', '').strip(),
            'categoria': request.form.get('categoria', 'Otro'),
            'favorito':  'Sí' if request.form.get('favorito') else 'No',
        }
        errores = _validar(datos['nombre'], datos['telefono'], datos['correo'])
        if not errores:
            _guardar_contacto(datos)
            flash('¡Contacto agregado correctamente!', 'se logro')
            return redirect(url_for('principal'))

    return render_template('agregar.html',
                           errores=errores,
                           categorias=CATEGORIAS)


# ── Editar contacto ────────────────────────────────────────
@app.route('/editar_contac', methods=['GET'])
@login_required
def editar_contac():
    contactos = _leer_contactos()
    busqueda  = request.args.get('busqueda', '').strip()
    if busqueda:
        contactos = [c for c in contactos
                     if busqueda.lower() in c['nombre'].lower()
                     or busqueda.lower() in c['apellido'].lower()]
    return render_template('editar_contac.html',
                           contactos=contactos,
                           busqueda=busqueda)


@app.route('/editar/<int:fila>', methods=['GET', 'POST'])
@login_required
def editar_formulario(fila):
    contactos = _leer_contactos()
    contacto  = next((c for c in contactos if c['fila'] == fila), None)
    if not contacto:
        flash('Contacto no encontrado.', 'error')
        return redirect(url_for('editar_contac'))

    errores = []
    if request.method == 'POST':
        datos = {
            'nombre':    request.form.get('nombre', '').strip(),
            'apellido':  request.form.get('apellido', '').strip(),
            'telefono':  request.form.get('telefono', '').strip(),
            'correo':    request.form.get('correo', '').strip(),
            'direccion': request.form.get('direccion', '').strip(),
            'categoria': request.form.get('categoria', 'Otro'),
            'favorito':  'Sí' if request.form.get('favorito') else 'No',
        }
        errores = _validar(datos['nombre'], datos['telefono'], datos['correo'])
        if not errores:
            _actualizar_contacto(fila, datos)
            flash('¡Contacto actualizado correctamente!', 'se logro')
            return redirect(url_for('principal'))

    return render_template('editar_formulario.html',
                           contacto=contacto,
                           errores=errores,
                           categorias=CATEGORIAS)


# ── Eliminar contacto ──────────────────────────────────────
@app.route('/eliminar/<int:fila>', methods=['POST'])
@login_required
def eliminar(fila):
    _eliminar_contacto(fila)
    flash('Contacto eliminado correctamente.', 'success')
    return redirect(url_for('principal'))


# ── Más información ──────────────────────────────
@app.route('/detalle/<int:fila>')
@login_required
def detalle(fila):
    contactos = _leer_contactos()
    contacto  = next((c for c in contactos if c['fila'] == fila), None)
    if not contacto:
        flash('Contacto no encontrado.', 'error')
        return redirect(url_for('principal'))
    return render_template('detalle.html', contacto=contacto)


# ── Reporte ────────────────────────────────────────────────
@app.route('/reporte_contac')
@login_required
def reporte_contac():
    contactos  = _leer_contactos()
    total      = len(contactos)
    favoritos  = sum(1 for c in contactos if c['favorito'] == 'Sí')
    por_categoria = {}
    for cat in CATEGORIAS:
        por_categoria[cat] = sum(1 for c in contactos if c['categoria'] == cat)
    return render_template('reporte_contac.html',
                           total=total,
                           favoritos=favoritos,
                           por_categoria=por_categoria)


# ── Cuenta usuario ──────────────────
@app.route('/cuenta_usuario')
@login_required
def cuenta_usuario():
    return render_template('cuenta_usuario.html')


# Final
if __name__ == '__main__':
    _init_excel()          
    app.run(debug=True)


